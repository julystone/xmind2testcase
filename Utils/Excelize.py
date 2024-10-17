import csv
import re
from typing import Any, List

import openpyxl
from openpyxl.styles import Alignment, Font
from pydantic import BaseModel, Field


class Case:
    def __repr__(self):
        return f"\n In __repr__：\n{repr(self.__dict__)}"


class SheetFoo(BaseModel):
    sheet_list: List[str] = Field(default_factory=list)
    selected_sheet: Any = None
    max_row: int = 0
    max_column: int = 0

    class Config:
        arbitrary_types_allowed = True

    def update_sheet_properties(self):
        """
        更新表格的最大行和最大列属性
        """
        if self.selected_sheet is not None:
            try:
                self.max_row = self.selected_sheet.max_row
                self.max_column = self.selected_sheet.max_column
            except AttributeError as e:
                print(f"更新表格属性时出现错误：{e}")


def csv_2_excel(csv_path: str, output_path: str = 'output.xlsx', hide_columns=None, colum_widths=None,
                style_dict=None) -> str:
    """
    将csv文件转换为excel文件
    :param csv_path: 输入的csv文件路径
    :param output_path: 输出的excel文件路径，默认为 output.xlsx
    :param hide_columns: 隐藏的列名列表
    :param colum_widths: 列宽字典，例如 {'A': 20, 'B': 30}
    :param style_dict: 单元格样式字典
    :return: 输出的excel文件路径
    """
    if style_dict is None:
        style_dict = {}
    if colum_widths is None:
        colum_widths = {}
    # try:
    wb = openpyxl.Workbook()
    ws = wb.active

    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    width_auto_fit(ws, colum_widths)
    use_style(ws, style_dict)
    column_hide(ws, hide_columns)

    wb.save(output_path)
    return output_path

    # except Exception as e:
    #     error_msg = f"转换过程中出现错误：{e}"
    #     print(error_msg)
    #     return error_msg


def column_hide(ws, column_name: str | List[str]):
    """
    隐藏指定的列
    :param ws: 工作表
    :param column_name: 列字母或字母列表
    """
    if isinstance(column_name, list):
        for column in column_name:
            ws.column_dimensions.group(column, hidden=True)
    else:
        ws.column_dimensions.group(column_name, hidden=True)


def width_auto_fit(ws, column_widths: dict):
    """
    自动调整列宽
    :param ws: 工作表
    :param column_widths: 列宽字典
    """

    for col_letter, width in column_widths.items():
        col_dimension = ws.column_dimensions[col_letter]
        col_dimension.width = width


def use_style(ws, style_dict: dict):
    """
    设置单元格样式
    :param ws: 工作表
    :param style_dict: 单元格样式字典
    """

    for column in ws.columns:
        for cell in column:
            cell.alignment = Alignment(**style_dict['alignment'])

    for cell in ws[1]:
        # cell.font = Font(**style_dict['font'])
        cell.font = Font(size=style_dict['font']['font_size'], bold=style_dict['font']['bold'],
                         color=style_dict['font']['font_color'])


def split_letter_and_number(s: str):
    """
    将字母和数字组合的字符串拆分成行和列
    :param s: 例如 "B12"
    :return: letter, number 例如: B, 12
    """
    match = re.match(r'(\D+)(\d+)', s)
    return (match.group(1), match.group(2)) if match else (None, None)


def get_row_column(string: str) -> (int, int):
    """
    将字母和数字组合的字符串拆分成行和列
    :param string: 例如 "B12"
    :return: row, column 例如: 12, 2
    """
    char, row = split_letter_and_number(string)
    column = sum((ord(letter.upper()) - ord('A') + 1) * (26 ** index) for index, letter in enumerate(char[::-1]))
    return int(row), column


class ReadExcel:
    """
    读取excel数据的类
    """

    def __init__(self, file_name: str, sheet_name: str = None, if_new_column: str = None):
        self.file_name = file_name
        self.if_new_column = if_new_column
        self.wb = openpyxl.load_workbook(self.file_name)
        self.sheet_name = sheet_name if sheet_name else self.wb.sheetnames[0]

        self.Sheet = SheetFoo(
            sheet_list=self.sheet_lists,
            selected_sheet=self.selected_sheet,
            max_row=self.selected_sheet.max_row,
            max_column=self.selected_sheet.max_column
        )

    def __del__(self):
        self.wb.close()

    @property
    def sheet_lists(self) -> List[str]:
        return self.wb.sheetnames

    @property
    def selected_sheet(self):
        return self.wb[self.sheet_name]

    @selected_sheet.setter
    def selected_sheet(self, sheet_name: str):
        if sheet_name not in self.sheet_lists:
            print("表单不存在")
            return
        self.sheet_name = sheet_name
        self.Sheet.selected_sheet = self.wb[sheet_name]
        self.Sheet.max_row, self.Sheet.max_column = self.max_row_column

    @property
    def max_row_column(self) -> (int, int):
        return self.Sheet.selected_sheet.max_row, self.Sheet.selected_sheet.max_column

    @property
    def latest_column_char(self) -> str:
        return chr(self.Sheet.max_column + ord('A'))

    @property
    def latest_row_num(self) -> str:
        return str(self.Sheet.max_row + 1)

    def w_data_origin(self, row: int, column: int, data: Any):
        self.Sheet.selected_sheet.cell(row, column, data)

    def w_data_char(self, string: str, data: Any):
        """
        根据列字母坐标写入数据
        :param string: 例如 “B12”
        :param data: 数据
        """
        self.w_data_origin(*get_row_column(string), data)

    def save(self):
        self.wb.save(self.file_name)

    def read_data_obj(self, sheet: str = None) -> List[Case]:
        """
        读取数据并返回用例对象列表
        :param sheet: 可选的工作表名
        :return: 包含用例对象的列表
        """
        if sheet:
            self.selected_sheet = sheet

        rows_data = list(self.Sheet.selected_sheet.rows)
        titles = [title.value for title in rows_data[0]]
        titles.append("max_column")

        if self.if_new_column and self.if_new_column not in titles:
            self.w_data_origin(1, self.Sheet.max_column + 1, self.if_new_column)

        cases = []
        for case in rows_data[1:]:
            case_obj = Case()
            data = [cell.value for cell in case]
            data.append(len(case))
            case_data = list(zip(titles, data))
            for title, value in case_data:
                if title == self.if_new_column or title is None:
                    continue
                setattr(case_obj, str(title), value)
            setattr(case_obj, 'row', case[0].row)
            cases.append(case_obj)

        return cases

    def hide_column(self, column_name: str | List[str]):
        """
        隐藏指定的列
        :param column_name: 列字母或字母列表
        """
        if isinstance(column_name, list):
            for column in column_name:
                self.Sheet.selected_sheet.column_dimensions.group(column, hidden=True)
        else:
            self.Sheet.selected_sheet.column_dimensions.group(column_name, hidden=True)


if __name__ == '__main__':
    r = ReadExcel("../SearchWords/SearchWords_v3.xlsx", sheet_name="July")
    r.selected_sheet = "Sheet1"
    res = get_row_column('ZZ100')
    print(res)
    r.save()
